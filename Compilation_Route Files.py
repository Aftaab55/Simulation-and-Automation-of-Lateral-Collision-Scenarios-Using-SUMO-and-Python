import openpyxl
from xml.etree import ElementTree as ET
import os

def extract_xml_data(xml_file):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        data = {
            'file_name': os.path.basename(xml_file),
            'actionStepLength': {},
            'minGapLat': {},
            'tau': {},
            'lcSigma': {}
        }
        
        for vtype in root.findall('vType'):
            vtype_id = vtype.get('id')
            if vtype_id in ['1', '2', '4']:
                data['actionStepLength'][vtype_id] = vtype.get('actionStepLength', 'N/A')
                data['minGapLat'][vtype_id] = vtype.get('minGapLat', 'N/A')
                data['tau'][vtype_id] = vtype.get('tau', 'N/A')
                data['lcSigma'][vtype_id] = vtype.get('lcSigma', 'N/A')
        
        return data
    except Exception as e:
        print(f"Error processing {xml_file}: {str(e)}")
        return None

def process_xml_files(input_dir, output_file):
    # Create or load workbook
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
    
    ws = wb.active
    
    # Write headers if new file
    if ws.max_row == 1:
        headers = [
            "File name",
            "actionStepLength_1", "actionStepLength_2", "actionStepLength_4",
            "minGapLat_1", "minGapLat_2", "minGapLat_4",
            "tau_1", "tau_2", "tau_4",
            "lcSigma_1", "lcSigma_2", "lcSigma_4"
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
    
    # Process each XML file
    for filename in os.listdir(input_dir):
        if filename.endswith('.xml'):
            xml_path = os.path.join(input_dir, filename)
            xml_data = extract_xml_data(xml_path)
            
            if xml_data:
                row = [
                    xml_data['file_name'],
                    xml_data['actionStepLength'].get('1', 'N/A'),
                    xml_data['actionStepLength'].get('2', 'N/A'),
                    xml_data['actionStepLength'].get('4', 'N/A'),
                    xml_data['minGapLat'].get('1', 'N/A'),
                    xml_data['minGapLat'].get('2', 'N/A'),
                    xml_data['minGapLat'].get('4', 'N/A'),
                    xml_data['tau'].get('1', 'N/A'),
                    xml_data['tau'].get('2', 'N/A'),
                    xml_data['tau'].get('4', 'N/A'),
                    xml_data['lcSigma'].get('1', 'N/A'),
                    xml_data['lcSigma'].get('2', 'N/A'),
                    xml_data['lcSigma'].get('4', 'N/A')
                ]
                ws.append(row)
    
    # Save the workbook
    wb.save(output_file)
    print(f"Processed {ws.max_row - 1} files. Output saved to {output_file}")

# Example usage
input_directory = r'C:\Users\aftaa\OneDrive\Desktop\Polito Mechanical\Thesis\Simulations\Automatisation\4\Output_new\Route_files'  # Current directory (where the script runs)
output_excel = r'C:\Users\aftaa\OneDrive\Desktop\Polito Mechanical\Thesis\Simulations\Automatisation\4\Output_new\Route_files\Compiled_Route_Data.xlsx'

process_xml_files(input_directory, output_excel)